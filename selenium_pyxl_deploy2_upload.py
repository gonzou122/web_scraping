import time
import openpyxl as pyxl
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

# Excelファイルの読み込み
file_path = 'C:/Users/1000228034/Desktop/scraping/heiretu/heiretu1/202104.xlsx'
wb = pyxl.load_workbook(file_path)
ws = wb.worksheets[0]

# headlessモードで設定
options = Options()
options.add_argument('--headless') 
options.add_argument('--disable-gpu') 
options.add_argument('--blink-settings=imagesEnabled=false')

# Webドライバーの読み込み
driver = webdriver.Chrome('C:/Users/1000228034/Desktop/scraping/heiretu/heiretu1/chromedriver.exe',chrome_options=options) # WebDriverのインスタンスを作成 pathを指定
driver.get('hogehoge') # URLを指定してブラウザを開く
time.sleep(1.5) # 2秒待機
# APMROのログイン情報の変数を入力
user_id = 'hogehoge'
okyakusama_bangou = 'hogehoge'
password = 'hogehoge'

search_box = driver.find_element(By.ID, value='prmUserNumber') # ユーザ名
search_box.send_keys(user_id) # ボックスに入力

time.sleep(0.25)

search_box = driver.find_element(By.ID, value='prmSiteNumber') # お客様番号
search_box.send_keys(okyakusama_bangou) # ボックスにパスワード入力

time.sleep(0.25) 

search_box = driver.find_element(By.ID, value='prmPassword') # お客様番号
search_box.send_keys(password) # ボックスにパスワード入力

time.sleep(0.25) 

search_box.submit()

time.sleep(1)

# APMROにログイン完了


#　100件の処理時間計測
start = time.perf_counter()

# ここから繰り返し処理 終了は終了番号の+1
# 一旦100件まで回す(max_rowは48239)
for idx in range(102 ,1001+1):
    
    # ここから型式検索
    if ws.cell(row=idx, column=4).value == 0:
        ws.cell(row=idx, column=6).value = '×'
        wb.save(file_path)
        driver.get('hogehoge')
    else:
        # iframeに移行
        iframe = driver.find_element(By.ID, value='exmg_search_form')
        driver.switch_to.frame(iframe)

        time.sleep(0.3)

            # カタログ購買サイトの検索欄に入力、検索ボタンをクリック
        search_box = driver.find_element(By.ID, value='keyword') 
        search_box.send_keys(ws.cell(row=idx, column=4).value)
        driver.find_element(By.ID, value='search_button').click()

        time.sleep(0.3)

        element_temp = driver.find_element(By.ID,value='hyo_list_content')

        if element_temp.text == '検索結果がありません。':
            ws.cell(row=idx, column=6).value = '×'
        else:
            ws.cell(row=idx, column=6).value = '〇'
            
        wb.save(file_path)
        driver.get('hogehoge')

    # ここから品名検索
    iframe = driver.find_element(By.ID, value='exmg_search_form')
    driver.switch_to.frame(iframe)

    time.sleep(0.3)

    # カタログ購買サイトの検索欄に入力、検索ボタンをクリック
    search_box = driver.find_element(By.ID, value='keyword') 
    search_box.send_keys(ws.cell(row=idx, column=5).value)
    driver.find_element(By.ID, value='search_button').click()

    time.sleep(0.3)

    element_temp = driver.find_element(By.ID,value='hyo_list_content')

    if element_temp.text == '検索結果がありません。':
        ws.cell(row=idx, column=7).value = '×'
    else:
        ws.cell(row=idx, column=7).value = '〇'
    
    wb.save(file_path)
    driver.get('hogehoge')
    print(f'{idx}行まで完了')

elapsed_time = time.perf_counter() - start
print(f'処理時間は{elapsed_time}秒です')

driver.quit() 
wb.save(file_path)